import argparse
import os
import re
from typing import Any, Dict, List, Tuple

import openpyxl


def norm(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", s)


def find_header_row(ws) -> Tuple[int, List[str]]:
    for r in range(1, min(ws.max_row, 50) + 1):
        vals = [c.value for c in ws[r]]
        non_empty = [v for v in vals if v is not None and str(v).strip() != ""]
        if len(non_empty) < 2:
            continue
        if any(re.search(r"[A-Za-z\u4e00-\u9fff]", str(v)) for v in non_empty):
            headers = [str(v).strip() if v is not None else "" for v in vals]
            return r, headers
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    return 1, headers


def row_to_dict(headers: List[str], row_vals: List[Any]) -> Dict[str, Any]:
    d = {}
    for i, h in enumerate(headers):
        key = h if h else f"col{i+1}"
        d[key] = row_vals[i] if i < len(row_vals) else None
    return d


def build_header_map(headers: List[str]) -> Dict[str, str]:
    mp = {}
    for h in headers:
        nh = norm(h)
        if nh and nh not in mp:
            mp[nh] = h
    return mp


def get_by_alias(row: Dict[str, Any], header_map: Dict[str, str], aliases: List[str]) -> Any:
    for a in aliases:
        na = norm(a)
        if na in header_map:
            return row.get(header_map[na])
    for na, orig in header_map.items():
        for a in aliases:
            if norm(a) and norm(a) in na:
                return row.get(orig)
    return None


def to_float(x: Any, default: float = 0.0) -> float:
    try:
        if x is None:
            return default
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if s == "":
            return default
        return float(s)
    except Exception:
        return default


def pct_up(new: float, old: float) -> str:
    if old <= 1e-12:
        return "N/A"
    return f"{(new - old) / old * 100:+.1f}%"


def pct_down(new: float, old: float) -> str:
    if old <= 1e-12:
        return "N/A"
    # down is good: (old-new)/old
    return f"{(old - new) / old * 100:+.1f}%"


def score_tuple(rec: float, top1: float, refusal: float, p95: float, avg: float):
    return (rec, top1, refusal, -p95, -avg)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default="reports/experiments.xlsx", help="Path to experiments.xlsx")
    ap.add_argument("--sheet", default="experiments", help="Sheet name")
    ap.add_argument("--baseline", default="", help="baseline run_name (optional). If empty: use first row")
    ap.add_argument("--best", default="", help="best run_name (optional). If empty: auto pick best")
    args = ap.parse_args()

    if not os.path.exists(args.excel):
        raise SystemExit(f"[ERR] Excel not found: {os.path.abspath(args.excel)}")

    wb = openpyxl.load_workbook(args.excel)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb[wb.sheetnames[0]]

    header_row, headers = find_header_row(ws)
    header_map = build_header_map(headers)

    rows: List[Dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [c.value for c in ws[r]]
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        rows.append(row_to_dict(headers, vals))

    if not rows:
        print("[ERR] No data rows found.")
        print("Headers detected:", headers)
        return

    def col(row, *aliases):
        return get_by_alias(row, header_map, list(aliases))

    # build enriched list
    enriched = []
    for row in rows:
        recall = to_float(col(row, "recall@5", "recall5", "recall_at_5", "recallk", "recall_at_k", "recall"), 0.0)
        top1 = to_float(col(row, "top1", "top1acc", "top1_accuracy", "top1@1"), 0.0)
        refusal = to_float(col(row, "refusalacc", "refusal_acc", "refusal_accuracy", "refusal"), 0.0)
        avg_ms = to_float(col(row, "avg_ms", "avgms", "avg_latency_ms", "mean_ms"), 0.0)
        p95_ms = to_float(col(row, "p95_ms", "p95ms", "p95_latency_ms"), 0.0)

        run_name = str(col(row, "run_name", "run", "name") or "")
        chunk = col(row, "chunk_chars", "chunk", "chunksize") or ""
        overlap = col(row, "overlap_chars", "overlap") or ""
        alpha = col(row, "alpha") or ""
        min_score = col(row, "min_score", "minscore") or ""
        report_path = str(col(row, "report_path", "report", "md_report") or "")

        enriched.append({
            "row": row,
            "run_name": run_name,
            "chunk": chunk,
            "overlap": overlap,
            "alpha": alpha,
            "min_score": min_score,
            "recall": recall,
            "top1": top1,
            "refusal": refusal,
            "avg_ms": avg_ms,
            "p95_ms": p95_ms,
            "report_path": report_path,
            "score": score_tuple(recall, top1, refusal, p95_ms, avg_ms)
        })

    # baseline
    baseline = None
    if args.baseline:
        for it in enriched:
            if it["run_name"] == args.baseline:
                baseline = it
                break
        if baseline is None:
            raise SystemExit(f"[ERR] baseline run_name not found: {args.baseline}")
    else:
        # use first row as baseline
        baseline = enriched[0]

    # best
    best = None
    if args.best:
        for it in enriched:
            if it["run_name"] == args.best:
                best = it
                break
        if best is None:
            raise SystemExit(f"[ERR] best run_name not found: {args.best}")
    else:
        best = sorted(enriched, key=lambda x: x["score"], reverse=True)[0]

    # compute deltas
    r_old, r_new = baseline["recall"], best["recall"]
    t_old, t_new = baseline["top1"], best["top1"]
    f_old, f_new = baseline["refusal"], best["refusal"]
    p_old, p_new = baseline["p95_ms"], best["p95_ms"]
    a_old, a_new = baseline["avg_ms"], best["avg_ms"]

    # bullets
    bullets = []

    bullets.append(
        "ProdTraceRAG：实现企业级可追溯 RAG 问答平台（FastAPI + Chroma 向量库 + BM25 混合检索），"
        "支持本地文档接入、证据引用(citations)、二阶段拒答/澄清、审计与参数可观测。"
    )

    bullets.append(
        f"构建评测闭环（≥50 QA，自动生成 Markdown 报告 + experiments.xlsx 试验台账），"
        f"通过网格搜索调参（chunk={best['chunk']}, α={best['alpha']}, min_score={best['min_score']}），"
        f"Recall@5 {r_old:.3f}→{r_new:.3f}（{pct_up(r_new, r_old)}），Top1 {t_old:.3f}→{t_new:.3f}（{pct_up(t_new, t_old)}），"
        f"拒答准确率 {f_old:.3f}→{f_new:.3f}（{pct_up(f_new, f_old)}）。"
    )

    bullets.append(
        f"在质量提升同时优化延迟：P95 {p_old:.0f}ms→{p_new:.0f}ms（{pct_down(p_new, p_old)}），"
        f"avg {a_old:.0f}ms→{a_new:.0f}ms（{pct_down(a_new, a_old)}）；"
        f"最佳配置报告：{best['report_path'] or 'reports/*.md'}。"
    )

    print("=== BASELINE ===")
    print(f"run_name: {baseline['run_name']}")
    print(f"R@5={r_old:.3f}, Top1={t_old:.3f}, RefusalAcc={f_old:.3f}, avg={a_old:.1f}ms, p95={p_old:.1f}ms\n")

    print("=== BEST ===")
    print(f"run_name: {best['run_name']}")
    print(f"chunk={best['chunk']}, alpha={best['alpha']}, min_score={best['min_score']}")
    print(f"R@5={r_new:.3f}, Top1={t_new:.3f}, RefusalAcc={f_new:.3f}, avg={a_new:.1f}ms, p95={p_new:.1f}ms\n")

    print("=== RESUME BULLETS ===")
    for b in bullets:
        print(f"- {b}")


if __name__ == "__main__":
    main()
