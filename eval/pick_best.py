import argparse
import os
import re
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


def norm(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    # keep alnum only
    return re.sub(r"[^a-z0-9]+", "", s)


def find_header_row(ws) -> Tuple[int, List[str]]:
    """
    Find the first row that looks like a header row.
    Strategy:
      - pick first row where >=2 cells are non-empty strings or contain letters.
    """
    for r in range(1, min(ws.max_row, 50) + 1):
        vals = [c.value for c in ws[r]]
        non_empty = [v for v in vals if v is not None and str(v).strip() != ""]
        if len(non_empty) < 2:
            continue
        # if at least one looks like a label
        if any(re.search(r"[A-Za-z\u4e00-\u9fff]", str(v)) for v in non_empty):
            headers = [str(v).strip() if v is not None else "" for v in vals]
            return r, headers
    # fallback: row1
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    return 1, headers


def row_to_dict(headers: List[str], row_vals: List[Any]) -> Dict[str, Any]:
    d = {}
    for i, h in enumerate(headers):
        key = h if h else f"col{i+1}"
        d[key] = row_vals[i] if i < len(row_vals) else None
    return d


def build_header_map(headers: List[str]) -> Dict[str, str]:
    """
    map normalized header -> original header
    """
    mp = {}
    for h in headers:
        nh = norm(h)
        if nh and nh not in mp:
            mp[nh] = h
    return mp


def get_by_alias(row: Dict[str, Any], header_map: Dict[str, str], aliases: List[str]) -> Any:
    """
    aliases are normalized candidates.
    """
    for a in aliases:
        na = norm(a)
        if na in header_map:
            return row.get(header_map[na])
    # fuzzy contains match
    for na, orig in header_map.items():
        for a in aliases:
            if norm(a) and norm(a) in na:
                return row.get(orig)
    return None


def to_float(x: Any, default: float = 0.0) -> float:
    try:
        if x is None:
            return default
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if s == "":
            return default
        return float(s)
    except Exception:
        return default


def to_int(x: Any, default: int = 0) -> int:
    try:
        if x is None:
            return default
        if isinstance(x, int):
            return x
        if isinstance(x, float):
            return int(x)
        s = str(x).strip()
        if s == "":
            return default
        return int(float(s))
    except Exception:
        return default


def score_tuple(rec: float, top1: float, refusal: float, p95: float, avg: float) -> Tuple[float, float, float, float, float]:
    return (rec, top1, refusal, -p95, -avg)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default="reports/experiments.xlsx", help="Path to experiments.xlsx")
    ap.add_argument("--sheet", default="experiments", help="Sheet name (default: experiments)")
    ap.add_argument("--topn", type=int, default=5, help="Show Top N runs")
    args = ap.parse_args()

    if not os.path.exists(args.excel):
        raise SystemExit(f"[ERR] Excel not found: {os.path.abspath(args.excel)}")

    wb = openpyxl.load_workbook(args.excel)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb[wb.sheetnames[0]]

    header_row, headers = find_header_row(ws)
    header_map = build_header_map(headers)

    # collect data rows
    rows: List[Dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [c.value for c in ws[r]]
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        rows.append(row_to_dict(headers, vals))

    if not rows:
        print("[ERR] No data rows found in Excel.")
        print("Headers detected:", headers)
        return

    # infer key columns via aliases
    def col(row, *aliases):
        return get_by_alias(row, header_map, list(aliases))

    scored = []
    for row in rows:
        recall = to_float(col(row, "recall@5", "recall5", "recall_at_5", "recallk", "recall_at_k", "recall"), 0.0)
        top1 = to_float(col(row, "top1", "top1acc", "top1_accuracy", "top1@1"), 0.0)
        refusal = to_float(col(row, "refusalacc", "refusal_acc", "refusal_accuracy", "refusal"), 0.0)
        avg_ms = to_float(col(row, "avg_ms", "avgms", "avg_latency_ms", "mean_ms"), 0.0)
        p95_ms = to_float(col(row, "p95_ms", "p95ms", "p95_latency_ms"), 0.0)

        run_name = col(row, "run_name", "run", "name") or ""
        chunk = col(row, "chunk_chars", "chunk", "chunksize") or ""
        overlap = col(row, "overlap_chars", "overlap") or ""
        alpha = col(row, "alpha") or ""
        min_score = col(row, "min_score", "minscore") or ""
        report_path = col(row, "report_path", "report", "md_report") or ""

        scored.append({
            "row": row,
            "run_name": str(run_name),
            "chunk": chunk,
            "overlap": overlap,
            "alpha": alpha,
            "min_score": min_score,
            "recall": recall,
            "top1": top1,
            "refusal": refusal,
            "avg_ms": avg_ms,
            "p95_ms": p95_ms,
            "report_path": str(report_path),
            "score": score_tuple(recall, top1, refusal, p95_ms, avg_ms)
        })

    scored.sort(key=lambda x: x["score"], reverse=True)

    best = scored[0]
    print("=== BEST ===")
    print(f"run_name      : {best['run_name']}")
    print(f"chunk/overlap : {best['chunk']}/{best['overlap']}")
    print(f"alpha         : {best['alpha']}")
    print(f"min_score     : {best['min_score']}")
    print(f"Recall@5      : {best['recall']:.3f}")
    print(f"Top1          : {best['top1']:.3f}")
    print(f"RefusalAcc    : {best['refusal']:.3f}")
    print(f"avg_ms / p95  : {best['avg_ms']:.1f} / {best['p95_ms']:.1f}")
    if best["report_path"]:
        print(f"report_path   : {best['report_path']}")

    print("\n=== TOP {} ===".format(min(args.topn, len(scored))))
    for i, it in enumerate(scored[: args.topn], 1):
        print(
            f"{i:>2}. {it['run_name']:<28} "
            f"R@5={it['recall']:.3f}  Top1={it['top1']:.3f}  Ref={it['refusal']:.3f}  "
            f"p95={it['p95_ms']:.0f}ms  chunk={it['chunk']}  a={it['alpha']}  ms={it['min_score']}"
        )

    # also write best to a small text file for downstream scripts
    out = os.path.join(os.path.dirname(args.excel), "best_run.txt")
    with open(out, "w", encoding="utf-8") as f:
        f.write(best["run_name"])
    print(f"\n[OK] best_run saved to: {out}")


if __name__ == "__main__":
    main()
